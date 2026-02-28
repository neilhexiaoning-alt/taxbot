#!/usr/bin/env python3
"""Generate an Excel reimbursement report from receipt data.

Usage:
    python generate_report.py <json_file> [--output <dir>] [--name <filename>]
    python generate_report.py --stdin --output <dir> [--name <filename>]

Arguments:
    json_file       Path to JSON file containing receipt data
    --stdin         Read JSON data from stdin instead of a file
    --output, -o    Output directory (default: same directory as json_file)
    --name, -n      Output filename without extension (default: 报销单_YYYYMMDD_HHMMSS)
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# -- Style constants ----------------------------------------------------------

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=16)
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=11, color="666666")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
TOTAL_FONT = Font(name="Microsoft YaHei", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
AMOUNT_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def apply_body_style(cell, alignment=LEFT):
    cell.font = BODY_FONT
    cell.alignment = alignment
    cell.border = THIN_BORDER


def apply_total_style(cell):
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER


# -- Sheet builders ------------------------------------------------------------

def build_detail_sheet(wb, data):
    """Build the receipt detail sheet."""
    ws = wb.active
    ws.title = "报销明细"

    title = data.get("title", "报销单")
    applicant = data.get("applicant", "")
    department = data.get("department", "")
    items = data.get("items", [])

    # Title row
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = title
    cell.font = TITLE_FONT
    cell.alignment = CENTER

    # Meta row
    meta_parts = []
    if applicant:
        meta_parts.append(f"报销人：{applicant}")
    if department:
        meta_parts.append(f"部门：{department}")
    meta_parts.append(f"制表日期：{datetime.now().strftime('%Y-%m-%d')}")
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "    ".join(meta_parts)
    cell.font = SUBTITLE_FONT
    cell.alignment = CENTER

    # Header row
    headers = ["序号", "日期", "费用类别", "摘要", "商家/卖方", "票据类型", "票据号码", "金额(元)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header_style(cell)

    # Data rows
    total_amount = 0.0
    for row_idx, item in enumerate(items, 5):
        seq = row_idx - 4
        amount = item.get("amount") or 0
        total_amount += float(amount)

        values = [
            seq,
            item.get("date", ""),
            item.get("category", "其他费用"),
            item.get("summary", ""),
            item.get("vendor", ""),
            item.get("receipt_type", ""),
            item.get("receipt_no", ""),
            float(amount),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                apply_body_style(cell, CENTER)
            elif col_idx == 8:
                apply_body_style(cell, RIGHT)
                cell.number_format = AMOUNT_FORMAT
            else:
                apply_body_style(cell)

    # Total row
    total_row = 5 + len(items)
    ws.merge_cells(f"A{total_row}:G{total_row}")
    cell = ws.cell(row=total_row, column=1, value="合  计")
    apply_total_style(cell)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(2, 8):
        c = ws.cell(row=total_row, column=col)
        apply_total_style(c)

    cell = ws.cell(row=total_row, column=8, value=total_amount)
    apply_total_style(cell)
    cell.number_format = AMOUNT_FORMAT
    cell.alignment = RIGHT

    # Column widths
    col_widths = [6, 14, 12, 30, 22, 16, 20, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    return total_amount


def build_summary_sheet(wb, data, total_amount):
    """Build the category summary sheet."""
    ws = wb.create_sheet("分类汇总")
    items = data.get("items", [])

    cat_stats = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for item in items:
        cat = item.get("category", "其他费用")
        cat_stats[cat]["count"] += 1
        cat_stats[cat]["amount"] += float(item.get("amount") or 0)

    sorted_cats = sorted(cat_stats.items(), key=lambda x: x[1]["amount"], reverse=True)

    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "费用分类汇总"
    cell.font = TITLE_FONT
    cell.alignment = CENTER

    # Header
    headers = ["费用类别", "笔数", "小计金额(元)", "占比"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        apply_header_style(cell)

    # Data
    for row_idx, (cat, stats) in enumerate(sorted_cats, 4):
        pct = stats["amount"] / total_amount if total_amount > 0 else 0

        ws.cell(row=row_idx, column=1, value=cat)
        apply_body_style(ws.cell(row=row_idx, column=1))

        ws.cell(row=row_idx, column=2, value=stats["count"])
        apply_body_style(ws.cell(row=row_idx, column=2), CENTER)

        cell = ws.cell(row=row_idx, column=3, value=stats["amount"])
        apply_body_style(cell, RIGHT)
        cell.number_format = AMOUNT_FORMAT

        cell = ws.cell(row=row_idx, column=4, value=pct)
        apply_body_style(cell, CENTER)
        cell.number_format = '0.0%'

    # Total row
    total_row = 4 + len(sorted_cats)
    cell = ws.cell(row=total_row, column=1, value="合  计")
    apply_total_style(cell)
    cell.alignment = Alignment(horizontal="right", vertical="center")

    cell = ws.cell(row=total_row, column=2, value=len(items))
    apply_total_style(cell)
    cell.alignment = CENTER

    cell = ws.cell(row=total_row, column=3, value=total_amount)
    apply_total_style(cell)
    cell.number_format = AMOUNT_FORMAT
    cell.alignment = RIGHT

    cell = ws.cell(row=total_row, column=4, value=1.0 if total_amount > 0 else 0)
    apply_total_style(cell)
    cell.number_format = '0.0%'
    cell.alignment = CENTER

    for i, w in enumerate([18, 10, 18, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"


# -- Main ---------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate Excel reimbursement report from receipt JSON data.")
    parser.add_argument("json_file", nargs="?", default=None, help="Path to JSON file with receipt data")
    parser.add_argument("--stdin", action="store_true", help="Read JSON data from stdin instead of a file")
    parser.add_argument("-o", "--output", default=None, help="Output directory (default: same as json_file)")
    parser.add_argument("-n", "--name", default=None, help="Output filename without extension")
    args = parser.parse_args()

    if args.stdin:
        raw = sys.stdin.read()
        data = json.loads(raw)
    elif args.json_file:
        with open(args.json_file, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
    else:
        print("Error: provide a json_file path or use --stdin", file=sys.stderr)
        sys.exit(1)

    output_dir = args.output or (os.path.dirname(os.path.abspath(args.json_file)) if args.json_file else os.getcwd())
    os.makedirs(output_dir, exist_ok=True)

    filename = args.name or f"报销单_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    output_path = os.path.join(output_dir, f"{filename}.xlsx")

    wb = Workbook()
    total_amount = build_detail_sheet(wb, data)
    build_summary_sheet(wb, data, total_amount)

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
