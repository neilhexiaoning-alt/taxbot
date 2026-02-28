#!/usr/bin/env python3
"""Extract and display text content from a single file.

Usage:
    python read_file.py <file_path> [--max-chars N]

Supports: PDF, DOCX, XLSX, TXT, MD, CSV, JSON, XML, HTML and more.
Outputs extracted text to stdout.

Exits with code 1 if file not found or unsupported format.
"""

import argparse
import os
import sys

DEFAULT_MAX_CHARS = 8000

TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".xml", ".html", ".htm", ".log", ".yaml", ".yml", ".ini", ".cfg", ".conf", ".bat", ".ps1", ".sh", ".py", ".js", ".ts"}


def _read_text_file(filepath, max_chars):
    """Read a plain text file with encoding fallback."""
    for enc in ["utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1"]:
        try:
            with open(filepath, "r", encoding=enc) as f:
                content = f.read(max_chars + 100)
            if len(content) > max_chars:
                content = content[:max_chars] + "\n...(truncated)"
            return content
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "[Error: unable to decode file with supported encodings]"


def _read_pdf(filepath, max_chars):
    """Extract text from PDF using available libraries."""
    for lib_name, reader_fn in [
        ("pdfplumber", _read_pdf_pdfplumber),
        ("PyPDF2", _read_pdf_pypdf2),
        ("pymupdf", _read_pdf_pymupdf),
    ]:
        try:
            result = reader_fn(filepath, max_chars)
            if result is not None:
                return result
        except ImportError:
            continue
        except Exception as e:
            return f"[{lib_name} error: {e}]"
    return "[no-pdf-library] Install pdfplumber: pip install pdfplumber"


def _read_pdf_pdfplumber(filepath, max_chars):
    import pdfplumber
    parts = []
    total = 0
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                parts.append(t.strip())
                total += len(t)
                if total > max_chars:
                    break
    text = "\n".join(parts)
    if not text:
        return None
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...(truncated)"
    return text


def _read_pdf_pypdf2(filepath, max_chars):
    from PyPDF2 import PdfReader
    reader = PdfReader(filepath)
    parts = []
    total = 0
    for page in reader.pages:
        t = page.extract_text()
        if t:
            parts.append(t.strip())
            total += len(t)
            if total > max_chars:
                break
    text = "\n".join(parts)
    if not text:
        return None
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...(truncated)"
    return text


def _read_pdf_pymupdf(filepath, max_chars):
    import fitz
    doc = fitz.open(filepath)
    parts = []
    total = 0
    for page in doc:
        t = page.get_text()
        if t:
            parts.append(t.strip())
            total += len(t)
            if total > max_chars:
                break
    text = "\n".join(parts)
    if not text:
        return None
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...(truncated)"
    return text


def _read_docx(filepath, max_chars):
    """Extract text from DOCX."""
    try:
        from docx import Document
    except ImportError:
        return "[needs python-docx] Install: pip install python-docx"
    doc = Document(filepath)
    parts = []
    total = 0
    for para in doc.paragraphs:
        parts.append(para.text)
        total += len(para.text)
        if total > max_chars:
            break
    text = "\n".join(parts)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...(truncated)"
    return text


def _read_xlsx(filepath, max_chars):
    """Extract text from XLSX."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "[needs openpyxl] Install: pip install openpyxl"
    wb = load_workbook(filepath, read_only=True, data_only=True)
    parts = []
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells)
            parts.append(line)
            total += len(line)
            if total > max_chars:
                break
        if total > max_chars:
            break
    wb.close()
    text = "\n".join(parts)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...(truncated)"
    return text


def main():
    parser = argparse.ArgumentParser(description="Extract text from a file.")
    parser.add_argument("file_path", help="Path to the file")
    parser.add_argument("--max-chars", type=int, default=DEFAULT_MAX_CHARS, help="Max chars to output")
    args = parser.parse_args()

    filepath = os.path.normpath(args.file_path)
    max_chars = args.max_chars

    if not os.path.isfile(filepath):
        print(f"Error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    ext = os.path.splitext(filepath)[1].lower()
    filename = os.path.basename(filepath)
    size = os.path.getsize(filepath)

    print(f"File: {filename} ({size} bytes)")
    print(f"{'=' * 60}")

    if ext in TEXT_EXTENSIONS:
        text = _read_text_file(filepath, max_chars)
    elif ext == ".pdf":
        text = _read_pdf(filepath, max_chars)
    elif ext in (".docx", ".doc"):
        text = _read_docx(filepath, max_chars)
    elif ext in (".xlsx", ".xls"):
        text = _read_xlsx(filepath, max_chars)
    elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"):
        text = "[image-file] This is an image file. Text extraction requires OCR."
    else:
        text = f"[unsupported-format] Cannot extract text from {ext} files."

    print(text)


if __name__ == "__main__":
    main()
